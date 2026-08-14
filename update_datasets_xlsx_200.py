# -*- coding: utf-8 -*-
"""Actualiza DATASETS.xlsx tras la ampliacion del corpus documental a 200 PDFs
(4 categorias nuevas: economia_fen, agroexportacion_norte, agronomia_hidrica_norte,
estudios_complementarios, 83 PDFs en total), ya documentadas en DATASETS.md
(2026-07-27) pero ausentes de la version Excel del catalogo.

Agrega 83 filas nuevas a las hojas "Catalogo" y "Usados" (prefijos de ID F/G/H/I,
libres de colision con los existentes C/A/M/E/EV/T/R) y actualiza los totales de
"Resumen". No toca "Recomendados" ni "Corpus ENFEN" (ya estaban correctos).

Uso: python update_datasets_xlsx_200.py
"""
from copy import copy

import openpyxl

XLSX = "DATASETS.xlsx"

# (categoria label, prefijo ID, carpeta) -> lista de (archivo, entidad/fuente, tema)
CATEGORIES = [
    (
        "Economia y politica del FEN",
        "F",
        "economia_fen",
        [
            ("APESEG_Fichas_tecnicas_productos_agroalimentarios_seguro_agrario.pdf", "APESEG", "Fichas tecnicas de productos agroalimentarios de seguro agrario"),
            ("APESEG_Impactos_FEN_economia_regional_Piura_Lambayeque_LaLibertad.pdf", "APESEG", "Impactos del FEN en la economia regional (Piura/Lambayeque/La Libertad)"),
            ("APE_WP97_Impacto_FEN_agroexportaciones.pdf", "APE (Working Paper 97)", "Impacto del FEN en las agroexportaciones"),
            ("BCRP_DT007_2024_morosidad_sector_agroexportador.pdf", "BCRP (Documento de Trabajo 007-2024)", "Morosidad del sector agroexportador"),
            ("BCRP_DT009_2021_efectos_economicos_cambio_climatico_Peru.pdf", "BCRP (Documento de Trabajo 009-2021)", "Efectos economicos del cambio climatico en el Peru"),
            ("BCRP_RI_jun2023_recuadro1_FEN_impacto_economia_peruana.pdf", "BCRP (Reporte de Inflacion jun-2023, recuadro 1)", "Impacto del FEN en la economia peruana"),
            ("BCRP_RI_mar2026_recuadro2_Nino_costero_2026_actividad_economica.pdf", "BCRP (Reporte de Inflacion mar-2026, recuadro 2)", "Nino costero 2026 y actividad economica"),
            ("BCRP_Reporte_Estabilidad_Financiera_mayo2024.pdf", "BCRP", "Reporte de Estabilidad Financiera, mayo 2024"),
            ("BCRP_Reporte_Estabilidad_Financiera_mayo2025.pdf", "BCRP", "Reporte de Estabilidad Financiera, mayo 2025"),
            ("BID_ENESA_Programa_manejo_riesgo_agropecuario_Peru.pdf", "BID", "Programa ENESA de manejo de riesgo agropecuario"),
            ("CAF_Lecciones_El_Nino_1997_1998_Peru.pdf", "CAF", "Lecciones de El Nino 1997-1998 en el Peru"),
            ("Congreso_Beneficios_incentivos_tributarios_incl_Nino_costero.pdf", "Congreso de la Republica", "Beneficios e incentivos tributarios (incl. Nino costero)"),
            ("Defensoria_Informe005_2018_seguimiento_intervenciones_post_FEN.pdf", "Defensoria del Pueblo (Informe 005-2018)", "Seguimiento de intervenciones post-FEN"),
            ("Defensoria_Informe178_concesiones_viales_lecciones_Nino_costero.pdf", "Defensoria del Pueblo (Informe 178)", "Concesiones viales: lecciones del Nino costero"),
            ("Dialnet_Derivados_financieros_efectos_FEN_BVL.pdf", "Dialnet / academico", "Derivados financieros y efectos del FEN en la BVL"),
            ("ENFEN_Informe_Tecnico_Extraordinario_001_2017_Nino_Costero.pdf", "Comite ENFEN", "Informe Tecnico Extraordinario 001-2017 (Nino costero)"),
            ("INDECI_Compendio_estadistico_2017_gestion_reactiva.pdf", "INDECI", "Compendio estadistico 2017 de gestion reactiva"),
            ("INDECI_Norma_complementaria_declaratoria_emergencia_SINAGERD.pdf", "INDECI", "Norma complementaria de declaratoria de emergencia (SINAGERD)"),
            ("INGEMMET_Evaluacion_geologica_Nino_costero_Tumbes_2017.pdf", "INGEMMET", "Evaluacion geologica del Nino costero en Tumbes 2017"),
            ("JICA_Estudio_danos_inundaciones_cuencas_Tumbes_Piura.pdf", "JICA", "Estudio de danos por inundaciones en cuencas de Tumbes y Piura"),
            ("KPMG_Riesgo_cambiario_Peru.pdf", "KPMG", "Riesgo cambiario en el Peru"),
            ("MIDAGRI_Exportaciones_agrarias_Peru_2023.pdf", "MIDAGRI", "Exportaciones agrarias del Peru 2023"),
            ("MIDAGRI_Informe5_Nino_costero_campanas_agricolas_2022_2024.pdf", "MIDAGRI (Informe 5)", "Nino costero y campanas agricolas 2022-2024"),
            ("MINAM_Dossier_El_Nino_en_el_Peru.pdf", "MINAM", "Dossier El Nino en el Peru"),
            ("MINAM_SINIA_Cambio_climatico_norte_Peru.pdf", "MINAM (SINIA)", "Cambio climatico en el norte del Peru"),
            ("OSITRAN_Resiliencia_carreteras_concesionadas_desastres_naturales.pdf", "OSITRAN", "Resiliencia de carreteras concesionadas ante desastres naturales"),
            ("PIRC_Plan_Integral_Reconstruccion_con_Cambios_set2017.pdf", "PIRC", "Plan Integral de Reconstruccion con Cambios (set-2017)"),
            ("PropuestaCiudadana_Balance_Reconstruccion_con_Cambios_2024.pdf", "Propuesta Ciudadana", "Balance de la Reconstruccion con Cambios 2024"),
            ("SENACE_IGAPRO_Evaluacion_ambiental_proyectos_PIRC.pdf", "SENACE (IGAPRO)", "Evaluacion ambiental de proyectos del PIRC"),
            ("ULima_Tesis_Nino_costero_2017_morosidad_crediticia.pdf", "Universidad de Lima (tesis)", "Nino costero 2017 y morosidad crediticia"),
            ("arXiv_2008_04887_consumer_behavior_El_Nino_2017_Peru.pdf", "arXiv", "Comportamiento del consumidor durante El Nino 2017 en el Peru"),
        ],
    ),
    (
        "Agroexportacion multicultivo",
        "G",
        "agroexportacion_norte",
        [
            ("ADEX_CIEN_Exportacion_Organicos_Diciembre2025.pdf", "ADEX / CIEN", "Exportacion de organicos, diciembre 2025"),
            ("ADEX_CIEN_Nota_Coyuntural_Mango_Enero2025.pdf", "ADEX / CIEN", "Nota coyuntural del mango, enero 2025"),
            ("ADEX_CIEN_Reporte_Exportaciones_Diciembre2025.pdf", "ADEX / CIEN", "Reporte de exportaciones, diciembre 2025"),
            ("BCRP_Reporte_Inflacion_junio2026.pdf", "BCRP", "Reporte de Inflacion, junio 2026"),
            ("CEPAL_Negocio_internacional_esparrago_Peru.pdf", "CEPAL", "Negocio internacional del esparrago peruano"),
            ("Camposol_Reporte_Sostenibilidad_2024.pdf", "Camposol", "Reporte de sostenibilidad 2024"),
            ("Exportemos_Requisitos_Fitosanitarios_Agroexportacion_2023.pdf", "Exportemos", "Requisitos fitosanitarios de agroexportacion 2023"),
            ("LACCEI_2024_Internal_logistics_agroexport_process.pdf", "LACCEI", "Logistica interna del proceso agroexportador"),
            ("LACCEI_2025_Peruvian_agroexport_competitiveness_sustainability.pdf", "LACCEI", "Competitividad y sostenibilidad de la agroexportacion peruana"),
            ("OPIP_RAM_Palta_Estados_Unidos.pdf", "OPIP", "Requisitos de acceso a mercado (RAM) de palta a EE.UU."),
            ("PROMPERU_Ficha_Mercado_EEUU_Uvas_Frescas_2020.pdf", "PROMPERU", "Ficha de mercado: uvas frescas a EE.UU. (2020)"),
            ("PROMPERU_Ficha_Mercado_Mango_abril2024.pdf", "PROMPERU", "Ficha de mercado del mango, abril 2024"),
            ("PROMPERU_Ficha_Mercado_Palta_febrero2025.pdf", "PROMPERU", "Ficha de mercado de la palta, febrero 2025"),
            ("PROMPERU_Ficha_Producto_Esparrago_fresco.pdf", "PROMPERU", "Ficha de producto: esparrago fresco"),
            ("PROMPERU_Informe_Mensual_Exportaciones_setiembre2025.pdf", "PROMPERU", "Informe mensual de exportaciones, setiembre 2025"),
            ("PROMPERU_Informe_Mercado_Arandanos_Estados_Unidos.pdf", "PROMPERU", "Informe de mercado: arandanos en Estados Unidos"),
            ("SENASA_Certificacion_Fitosanitaria_Mango.pdf", "SENASA", "Certificacion fitosanitaria del mango"),
            ("SENASA_Certificacion_Fitosanitaria_Palto_Hass.pdf", "SENASA", "Certificacion fitosanitaria de palto Hass"),
            ("SENASA_Certificacion_Fitosanitaria_Supervision_Esparrago.pdf", "SENASA", "Certificacion fitosanitaria y supervision del esparrago"),
            ("SENASA_Certificacion_Fitosanitaria_Uva_Fresca.pdf", "SENASA", "Certificacion fitosanitaria de uva fresca"),
            ("SENASA_Procedimiento_Exportacion_Uva.pdf", "SENASA", "Procedimiento de exportacion de uva"),
            ("SENASA_Procedimiento_Integrado_Exportacion_Vegetal.pdf", "SENASA", "Procedimiento integrado de exportacion vegetal"),
        ],
    ),
    (
        "Agronomia/Hidrica complementaria",
        "H",
        "agronomia_hidrica_norte",
        [
            ("ANA_Politica_Estrategia_Nacional_Recursos_Hidricos.pdf", "Autoridad Nacional del Agua", "Gestion hidrica nacional"),
            ("BlueberriesConsulting_Manejo_Cosecha_Poscosecha_Arandano.pdf", "Blueberries Consulting", "Poscosecha y calidad del arandano"),
            ("BlueberriesConsulting_Manejo_Integrado_Plagas_Enfermedades_Arandano.pdf", "Blueberries Consulting", "Manejo integrado de plagas y enfermedades del arandano"),
            ("CEPES_PropuestaCiudadana_Asociatividad_Agricultura_Familiar_Ley31335.pdf", "CEPES / Propuesta Ciudadana", "Asociatividad y cooperativismo (Ley 31335)"),
            ("Congreso_Ley32434_Transformacion_Sector_Agrario_2025.pdf", "Congreso de la Republica", "Marco legal agrario vigente (Ley 32434, set-2025)"),
            ("PREDES_Manual_Riego_por_Goteo.pdf", "PREDES", "Riego tecnificado por goteo"),
            ("PropuestaCiudadana_FAEAGRO_problema_financiamiento.pdf", "Propuesta Ciudadana", "Financiamiento MYPE agrario"),
        ],
    ),
    (
        "Estudios complementarios",
        "I",
        "estudios_complementarios",
        [
            ("ADEX_CIEN_Reporte_Exportaciones_Enero2025.pdf", "ADEX / CIEN", "Reporte de exportaciones, enero 2025"),
            ("ANA_Plan_Gestion_Recursos_Hidricos_Cuenca_Chira_Piura.pdf", "Autoridad Nacional del Agua", "Plan de gestion de recursos hidricos, cuenca Chira-Piura"),
            ("BBVA_Research_El_Nino_America_Latina_2026.pdf", "BBVA Research", "El Nino en America Latina 2026"),
            ("BCRP_Sintesis_Actividad_Economica_Lambayeque_Mayo2025.pdf", "BCRP", "Sintesis de actividad economica de Lambayeque, mayo 2025"),
            ("CEPES_Mujeres_en_la_Agricultura_Peru_2024.pdf", "CEPES", "Mujeres en la agricultura del Peru 2024"),
            ("Camposol_Reporte_Sostenibilidad_2023.pdf", "Camposol", "Reporte de sostenibilidad 2023"),
            ("Datum_Tendencias_Consumo_Peru_Agosto2024.pdf", "Datum", "Tendencias de consumo en el Peru, agosto 2024"),
            ("Defensoria_Informe_Riesgos_Abastecimiento_Agua_2025.pdf", "Defensoria del Pueblo", "Riesgos de abastecimiento de agua 2025"),
            ("Exportemos_Guia_Mercado_HongKong.pdf", "Exportemos / PROMPERU", "Guia de mercado: Hong Kong"),
            ("FOGASA_Directiva_Seguro_Indexado_Alpacas.pdf", "FOGASA / MIDAGRI", "Directiva de seguro indexado para alpacas"),
            ("GlobalGAP_Reglamento_General_v6_Productores_Individuales.pdf", "GlobalGAP", "Reglamento general v6, productores individuales"),
            ("Hortifrut_Reporte_Sostenibilidad_2021.pdf", "Hortifrut", "Reporte de sostenibilidad 2021"),
            ("INDECI_Reporte_Lluvias_Intensas_Piura_Junio2026.pdf", "INDECI", "Reporte de lluvias intensas en Piura, junio 2026"),
            ("Ipsos_El_Consumidor_Peru_2024.pdf", "Ipsos", "El consumidor peruano 2024"),
            ("MINCETUR_PENX_2025_Plan_Estrategico_Nacional_Exportador.pdf", "MINCETUR", "Plan Estrategico Nacional Exportador (PENX) 2025"),
            ("MinCultura_Agroexportacion_Empleo_Genero_Peru.pdf", "Ministerio de Cultura", "Agroexportacion, empleo y genero en el Peru"),
            ("RainforestAlliance_Estandar_Agricultura_Sostenible_2020.pdf", "Rainforest Alliance", "Estandar de agricultura sostenible 2020"),
            ("SENASA_DS016_Reglamento_Registro_Control_Plaguicidas.pdf", "SENASA (D.S. 016)", "Reglamento de registro y control de plaguicidas"),
            ("SENASA_Manual_Sistema_Nacional_Vigilancia_Moscas_Fruta.pdf", "SENASA", "Manual del Sistema Nacional de Vigilancia de Moscas de la Fruta"),
            ("UAP_Tesis_Proyecto_Exportacion_Arandanos_EstadosUnidos.pdf", "UAP (tesis)", "Proyecto de exportacion de arandanos a Estados Unidos"),
            ("UCSM_Tesis_Comparativo_Exportacion_Arandano_EEUU_PaisesBajos.pdf", "UCSM (tesis)", "Comparativo de exportacion de arandano: EE.UU. vs. Paises Bajos"),
            ("ULaSalle_Zavalaga_Tesis_Mercado_Exportacion_Arandano.pdf", "Universidad La Salle (tesis Zavalaga)", "Mercado de exportacion del arandano"),
            ("WorldBank_Country_Climate_Development_Report_Peru_2022.pdf", "Banco Mundial", "Country Climate and Development Report, Peru 2022"),
        ],
    ),
]

TOTAL_NEW = sum(len(v[3]) for v in CATEGORIES)
assert TOTAL_NEW == 83, TOTAL_NEW

wb = openpyxl.load_workbook(XLSX)


def append_rows(sheet_name):
    ws = wb[sheet_name]
    template_row = 2  # estilo de referencia (fila de datos existente)
    styles = [copy(ws.cell(row=template_row, column=c).font) for c in range(1, 11)]
    fills = [copy(ws.cell(row=template_row, column=c).fill) for c in range(1, 11)]
    borders = [copy(ws.cell(row=template_row, column=c).border) for c in range(1, 11)]
    aligns = [copy(ws.cell(row=template_row, column=c).alignment) for c in range(1, 11)]
    row_height = ws.row_dimensions[template_row].height

    r = ws.max_row + 1
    count = 0
    for categoria, prefix, _folder, docs in CATEGORIES:
        for i, (fname, entidad, tema) in enumerate(docs, start=1):
            values = [
                f"{prefix}-{i:02d}",
                categoria,
                fname,
                entidad,
                tema,
                "Documento",
                "PDF",
                "Usado",
                "Corpus RAG",
                "Publica",
            ]
            for c, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = styles[c - 1]
                cell.fill = fills[c - 1]
                cell.border = borders[c - 1]
                cell.alignment = aligns[c - 1]
            if row_height:
                ws.row_dimensions[r].height = row_height
            r += 1
            count += 1
    return count


n1 = append_rows("Catálogo")
n2 = append_rows("Usados")
assert n1 == 83 and n2 == 83, (n1, n2)

# --- Resumen: totales + desglose de las 4 categorias nuevas ---
ws_r = wb["Resumen"]
totals = {c[0].value: c for c in ws_r.iter_rows(min_row=2, max_row=ws_r.max_row)}


def set_metric(label, value):
    for row in ws_r.iter_rows(min_row=2, max_row=ws_r.max_row):
        if row[0].value == label:
            row[1].value = value
            return
    raise KeyError(label)


set_metric("Total de entradas catalogadas", 135 + 83)
set_metric("Datasets/fuentes USADAS", 119 + 83)
set_metric("Total PDFs corpus RAG", 200)

# nuevas filas de desglose, insertadas antes de "Pares Q&A en golden dataset"
insert_before = None
for idx, row in enumerate(ws_r.iter_rows(min_row=2, max_row=ws_r.max_row), start=2):
    if row[0].value == "Pares Q&A en golden dataset":
        insert_before = idx
        break
assert insert_before is not None

new_metric_rows = [(categoria, len(docs)) for categoria, _p, _f, docs in CATEGORIES]
ws_r.insert_rows(insert_before, amount=len(new_metric_rows))

templ_font = copy(ws_r.cell(row=2, column=1).font)
templ_font_b = copy(ws_r.cell(row=2, column=2).font)
templ_align = copy(ws_r.cell(row=2, column=1).alignment)

for i, (label, value) in enumerate(new_metric_rows):
    rr = insert_before + i
    a = ws_r.cell(row=rr, column=1, value=label)
    b = ws_r.cell(row=rr, column=2, value=value)
    a.font = templ_font
    b.font = templ_font_b
    a.alignment = templ_align
    b.alignment = templ_align

wb.save(XLSX)
print(f"OK: {n1} filas nuevas en Catalogo, {n2} en Usados; Resumen actualizado "
      f"(218 entradas catalogadas, 202 usadas, 200 PDFs corpus RAG, "
      f"+{len(new_metric_rows)} filas de desglose por categoria nueva)")
