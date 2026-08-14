# -*- coding: utf-8 -*-
"""Segunda pasada sobre DATASETS.xlsx (tras update_datasets_xlsx_200.py), a pedido
del usuario: faltaban (1) los golden datasets de evaluacion reales (solo estaba el
de 8 pares, no los de 1000/150 que documenta DATASETS.md 1.2), (2) las fuentes NOAA
que el agente SI consume en produccion (CDO API v2 y CoastWatch ERDDAP; solo estaba
el indice ONI como "recomendado", no las integraciones reales) con sus URLs de
descarga, y (3) una columna de URL/Enlace en general (no existia ninguna).

Agrega columna K "URL / Enlace" a Catalogo, Usados, Recomendados y Corpus ENFEN;
agrega filas EV-02/EV-03 (golden dataset 1000 / muestra 150) y N-01/N-02 (NOAA CDO
API, NOAA CoastWatch ERDDAP) a Catalogo+Usados; puebla la URL en las 98 filas de
ENFEN (portal oficial IMARPE, verificado en vivo) y en EV-01/EV-02/EV-03 (ruta del
repo). Los 16 datasets "Recomendados" se completan aparte (requieren verificacion
web por fuente antes de citarlos).

Uso: python update_datasets_xlsx_urls_eval.py
"""
from copy import copy

import openpyxl

XLSX = "DATASETS.xlsx"

ENFEN_PORTAL = "https://enfen.imarpe.gob.pe/informes-tecnicos/"
NOAA_CDO_URL = "https://www.ncei.noaa.gov/cdo-web/api/v2"
NOAA_ERDDAP_URL = "https://coastwatch.noaa.gov/erddap/griddap/noaacwBLENDEDCsstDaily.json"

NEW_EVAL_ROWS = [
    (
        "EV-02", "Evaluacion", "golden_dataset_1000.jsonl", "Propio (curado)",
        "1000 pares pregunta/ground_truth (649 de informes ENFEN + 351 de "
        "documentos tematicos) - dataset por defecto de RAGAS + DeepEval",
        "Dataset Q&A", "JSONL", "Usado", "Evaluacion", "Propia",
        "evaluation/golden_dataset_1000.jsonl",
    ),
    (
        "EV-03", "Evaluacion", "eval_sample_150.jsonl", "Propio (derivado)",
        "Muestra estratificada por categoria (150 de 1000) para comparaciones "
        "rapidas antes/despues sin correr el set completo",
        "Dataset Q&A", "JSONL", "Usado", "Evaluacion", "Propia",
        "evaluation/eval_sample_150.jsonl",
    ),
]

NEW_NOAA_ROWS = [
    (
        "N-01", "Clima (NOAA)", "NOAA Climate Data Online API v2 (NCEI)", "NOAA/NCEI",
        "Precipitacion reciente por estacion GHCND mas cercana a coordenadas del "
        "norte peruano - senal temprana adicional del FEN (tool consultar_datos_noaa)",
        "API", "JSON", "Usado", "Tool del agente", "API (token)",
        NOAA_CDO_URL,
    ),
    (
        "N-02", "Clima (NOAA)", "NOAA CoastWatch ERDDAP - Blended SST diaria", "NOAA/CoastWatch",
        "SST diaria de puntos oceanicos del norte peruano - portal web-sst-monitor "
        "y alerta de anomalia SST (sst_alert.py, region Nino 1+2)",
        "API", "JSON (griddap)", "Usado", "Portal SST / alerta proactiva", "Abierta",
        NOAA_ERDDAP_URL,
    ),
]

wb = openpyxl.load_workbook(XLSX)

HEADER_LABEL = "URL / Enlace"


def add_url_column(sheet_name):
    ws = wb[sheet_name]
    col = ws.max_column + 1
    header_src = ws.cell(row=1, column=1)
    hcell = ws.cell(row=1, column=col, value=HEADER_LABEL)
    hcell.font = copy(header_src.font)
    hcell.fill = copy(header_src.fill)
    hcell.border = copy(header_src.border)
    hcell.alignment = copy(header_src.alignment)
    ws.column_dimensions[hcell.column_letter].width = 46
    return col


def style_data_cell(ws, r, col):
    cell = ws.cell(row=r, column=col)
    cell.font = copy(ws.cell(row=r, column=1).font)
    cell.fill = copy(ws.cell(row=r, column=1).fill)
    cell.border = copy(ws.cell(row=r, column=1).border)
    cell.alignment = copy(ws.cell(row=r, column=1).alignment)
    return cell


def append_full_rows(sheet_name, rows, url_col):
    ws = wb[sheet_name]
    template_row = 2
    styles = [copy(ws.cell(row=template_row, column=c).font) for c in range(1, 11)]
    fills = [copy(ws.cell(row=template_row, column=c).fill) for c in range(1, 11)]
    borders = [copy(ws.cell(row=template_row, column=c).border) for c in range(1, 11)]
    aligns = [copy(ws.cell(row=template_row, column=c).alignment) for c in range(1, 11)]
    row_height = ws.row_dimensions[template_row].height

    r = ws.max_row + 1
    for values in rows:
        data10, url = values[:10], values[10]
        for c, val in enumerate(data10, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = styles[c - 1]
            cell.fill = fills[c - 1]
            cell.border = borders[c - 1]
            cell.alignment = aligns[c - 1]
        ucell = ws.cell(row=r, column=url_col, value=url)
        ucell.font = styles[0]
        ucell.fill = fills[0]
        ucell.border = borders[0]
        ucell.alignment = aligns[0]
        if row_height:
            ws.row_dimensions[r].height = row_height
        r += 1


# 1) columna URL en las 4 hojas
url_cols = {}
for sheet in ["Catálogo", "Usados", "Recomendados", "Corpus ENFEN"]:
    url_cols[sheet] = add_url_column(sheet)

# 2) filas nuevas: golden datasets 1000/150 + fuentes NOAA reales
append_full_rows("Catálogo", NEW_EVAL_ROWS + NEW_NOAA_ROWS, url_cols["Catálogo"])
append_full_rows("Usados", NEW_EVAL_ROWS + NEW_NOAA_ROWS, url_cols["Usados"])

# 3) URL para EV-01 (fila ya existente) en Catalogo y Usados
for sheet in ["Catálogo", "Usados"]:
    ws = wb[sheet]
    col = url_cols[sheet]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value == "EV-01":
            style_data_cell(ws, row[0].row, col)
            ws.cell(row=row[0].row, column=col, value="evaluation/golden_dataset.jsonl")

# 4) URL del portal ENFEN en las 98 filas de "Corpus ENFEN" y en las E- de Catalogo/Usados
ws = wb["Corpus ENFEN"]
col = url_cols["Corpus ENFEN"]
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    style_data_cell(ws, row[0].row, col)
    ws.cell(row=row[0].row, column=col, value=ENFEN_PORTAL)

for sheet in ["Catálogo", "Usados"]:
    ws = wb[sheet]
    col = url_cols[sheet]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if str(row[0].value).startswith("E-"):
            style_data_cell(ws, row[0].row, col)
            ws.cell(row=row[0].row, column=col, value=ENFEN_PORTAL)

# --- Resumen: reflejar el dataset de evaluacion por defecto (1000, no 8) ---
ws_r = wb["Resumen"]
for row in ws_r.iter_rows(min_row=2, max_row=ws_r.max_row):
    if row[0].value == "Pares Q&A en golden dataset":
        row[0].value = "Pares Q&A en golden dataset (por defecto RAGAS/DeepEval)"
        row[1].value = 1000
        break

# nueva fila: total de entradas tras esta segunda pasada
for row in ws_r.iter_rows(min_row=2, max_row=ws_r.max_row):
    if row[0].value == "Total de entradas catalogadas":
        row[1].value = row[1].value + len(NEW_EVAL_ROWS) + len(NEW_NOAA_ROWS)
    if row[0].value == "Datasets/fuentes USADAS":
        row[1].value = row[1].value + len(NEW_EVAL_ROWS) + len(NEW_NOAA_ROWS)

wb.save(XLSX)
print("OK: columna 'URL / Enlace' agregada a Catalogo/Usados/Recomendados/Corpus ENFEN; "
      "+2 filas EV (golden dataset 1000/150), +2 filas N (NOAA CDO API / ERDDAP); "
      "98 filas ENFEN con URL del portal oficial IMARPE; Resumen actualizado "
      "(golden dataset por defecto = 1000).")
