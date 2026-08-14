# -*- coding: utf-8 -*-
"""Tercera pasada sobre DATASETS.xlsx: puebla la columna "URL / Enlace" (agregada
en update_datasets_xlsx_urls_eval.py) para las 16 filas R-01..R-16 ("Recomendados"),
en la hoja Recomendados y en su espejo dentro de Catalogo. URLs verificadas por
investigacion web puntual (no adivinadas) - ver notas de confianza en el commit/PR.

Dos correcciones factuales encontradas durante la verificacion, aplicadas como nota
breve en la columna Tema/Contenido (no se borra ni se cambia el estado Recomendado,
es solo contexto adicional):
- R-10 (indice ONI): NOAA reemplazo ONI por RONI como indice oficial ENSO en
  feb-2026; la tabla ONI se sigue publicando mensualmente pero ya no es el indice
  primario oficial.
- R-13 (golden dataset ampliado 50-100 preg./categoria): la recomendacion ya esta
  superada en la practica por EV-02 (golden_dataset_1000.jsonl, 1000 pares), que
  se agrego en la pasada anterior - se deja igual el estado "Recomendado" porque
  cambiar el estado es una decision de contenido, no un dato factual verificable.

Uso: python update_datasets_xlsx_recomendados_urls.py
"""
import openpyxl

XLSX = "DATASETS.xlsx"

URLS = {
    "R-01": "https://www.datosabiertos.gob.pe/ (portal general - sin deep-link confirmado para la subpartida 0810.40 especificamente)",
    "R-02": "https://www.siicex.gob.pe/",
    "R-03": "https://comtradeplus.un.org/",
    "R-04": "https://www.trademap.org/",
    "R-05": "https://www.fao.org/faostat/en/#data/QCL",
    "R-06": "https://siea.midagri.gob.pe/",
    "R-07": "https://www.gob.pe/institucion/senasa/colecciones/2518-lista-de-empacadoras-y-lugares-de-produccion-certificados-por-el-senasa-para-poder-exportar-desde-el-peru",
    "R-08": "https://www.datosabiertos.gob.pe/dataset/censo-nacional-agropecuario-cenagro-2012-instituto-nacional-de-estad%C3%ADstica-e-inform%C3%A1tica",
    "R-09": "https://www.senamhi.gob.pe/?p=descarga-datos-hidrometeorologicos",
    "R-10": "https://www.cpc.ncep.noaa.gov/products/analysis_monitoring/ensostuff/ONI_v5.php",
    "R-11": "https://estadisticas.bcrp.gob.pe/estadisticas/series/",
    "R-12": "https://open-meteo.com/en/docs · https://power.larc.nasa.gov/",
    "R-13": "",  # recurso propio, sin URL externa
    "R-14": "https://huggingface.co/datasets/miracl/miracl",
    "R-15": "https://huggingface.co/datasets/google/xquad · https://huggingface.co/datasets/facebook/mlqa",
    "R-16": "https://github.com/NVIDIA/garak · https://github.com/lakeraai/pint-benchmark",
}

TEMA_NOTES = {
    "R-10": (" [nota: NOAA reemplazo ONI por RONI como indice oficial ENSO en "
             "feb-2026; la tabla ONI se sigue publicando mensualmente]"),
    "R-13": " [nota: ya cubierto en la practica por EV-02, golden_dataset_1000.jsonl]",
}

wb = openpyxl.load_workbook(XLSX)


def apply(sheet_name):
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    id_col = header.index("ID") + 1
    tema_col = header.index("Tema / Contenido") + 1
    url_col = header.index("URL / Enlace") + 1
    n = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        rid = row[id_col - 1].value
        if rid in URLS:
            ws.cell(row=row[0].row, column=url_col, value=URLS[rid])
            if rid in TEMA_NOTES:
                cur = row[tema_col - 1].value or ""
                if TEMA_NOTES[rid] not in cur:
                    ws.cell(row=row[0].row, column=tema_col, value=cur + TEMA_NOTES[rid])
            n += 1
    return n


n1 = apply("Recomendados")
n2 = apply("Catálogo")
assert n1 == 16, n1
assert n2 == 16, n2

wb.save(XLSX)
print(f"OK: URL poblada en {n1} filas de Recomendados y {n2} filas espejo en Catalogo "
      f"(R-01..R-16); notas factuales agregadas en R-10 (ONI/RONI) y R-13 (superado por EV-02).")
